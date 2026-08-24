from __future__ import annotations

import json
import shutil
from pathlib import Path

from openpyxl import Workbook

from proofmesh.service import DocumentAuditService
from scripts.create_demo_bundle import _write_minimal_pdf


def test_scanned_like_pdf_is_reported_as_partial(tmp_path: Path) -> None:
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    _write_minimal_pdf(bundle / "scan.pdf", "")
    project_root = Path(__file__).resolve().parents[1]

    result = DocumentAuditService(project_root, tmp_path / "home").audit(str(bundle))
    run_dir = Path(result["run_dir"])
    file_results = json.loads((run_dir / "file_results.json").read_text(encoding="utf-8"))

    assert result["status"] == "partial"
    assert file_results == [
        {"relative_path": "scan.pdf", "status": "needs_ocr", "message": "没有读到文本层，需要 OCR。"}
    ]
    assert "部分检查没有完成" in (run_dir / "report.md").read_text(encoding="utf-8")
    assert "没有读到文本层，需要 OCR" in (run_dir / "report.html").read_text(encoding="utf-8")


def test_formula_without_cached_value_is_visible(tmp_path: Path) -> None:
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["指标", "数值"])
    sheet.append(["项目预算", "=100+26"])
    workbook.save(bundle / "公式.xlsx")
    project_root = Path(__file__).resolve().parents[1]

    result = DocumentAuditService(project_root, tmp_path / "home").audit(str(bundle))
    run_dir = Path(result["run_dir"])
    file_results = json.loads((run_dir / "file_results.json").read_text(encoding="utf-8"))

    assert result["status"] == "partial"
    assert file_results[0]["status"] == "parsed_with_warnings"
    assert "公式没有缓存结果" in file_results[0]["message"]


def test_different_column_periods_do_not_create_false_conflict(tmp_path: Path) -> None:
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    for name, year, value in [("历史.xlsx", "2025", 100), ("计划.xlsx", "2026", 200)]:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["指标", year])
        sheet.append(["项目预算", value])
        workbook.save(bundle / name)
    project_root = Path(__file__).resolve().parents[1]

    result = DocumentAuditService(project_root, tmp_path / "home").audit(str(bundle))
    issues = json.loads((Path(result["run_dir"]) / "issues.json").read_text(encoding="utf-8"))

    assert issues == []


def test_corrupt_workbook_is_not_reported_as_success(tmp_path: Path) -> None:
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    (bundle / "损坏.xlsx").write_bytes(b"not-an-xlsx")
    project_root = Path(__file__).resolve().parents[1]

    result = DocumentAuditService(project_root, tmp_path / "home").audit(str(bundle))
    run = json.loads((Path(result["run_dir"]) / "run.json").read_text(encoding="utf-8"))

    assert result["status"] == "partial"
    assert run["checked_file_count"] == 0
    assert len(run["errors"]) == 1


def test_missing_openvino_model_marks_run_partial(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    fake_root = tmp_path / "fake_project"
    (fake_root / "rules").mkdir(parents=True)
    shutil.copy2(project_root / "rules" / "default.yaml", fake_root / "rules" / "default.yaml")
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    workbook = Workbook()
    workbook.active.append(["项目预算", "126万元"])
    workbook.save(bundle / "预算.xlsx")

    result = DocumentAuditService(fake_root, tmp_path / "home").audit(str(bundle))
    run_dir = Path(result["run_dir"])
    model_info = json.loads((run_dir / "model_info.json").read_text(encoding="utf-8"))

    assert result["status"] == "partial"
    assert model_info["available"] is False
    assert "近似指标候选未执行" in (run_dir / "report.md").read_text(encoding="utf-8")


def test_invalid_model_manifest_marks_run_partial(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    fake_root = tmp_path / "fake_project"
    (fake_root / "rules").mkdir(parents=True)
    shutil.copy2(project_root / "rules" / "default.yaml", fake_root / "rules" / "default.yaml")
    model_dir = fake_root / "models" / "bge-small-zh-v1.5-openvino"
    model_dir.mkdir(parents=True)
    (model_dir / "model-manifest.json").write_text("{broken", encoding="utf-8")
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    workbook = Workbook()
    workbook.active.append(["项目预算", "126万元"])
    workbook.save(bundle / "预算.xlsx")

    result = DocumentAuditService(fake_root, tmp_path / "home").audit(str(bundle))

    assert result["status"] == "partial"
    assert result["agent_summary"]["incomplete_check_count"] == 1


def test_sheet_money_unit_does_not_override_percentage_format(tmp_path: Path) -> None:
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["单位：万元", None])
    sheet.append(["预算完成率", 0.5])
    sheet["B2"].number_format = "0%"
    workbook.save(bundle / "进度.xlsx")
    project_root = Path(__file__).resolve().parents[1]

    result = DocumentAuditService(project_root, tmp_path / "home").audit(str(bundle))
    observations = [
        json.loads(line)
        for line in (Path(result["run_dir"]) / "fact_observations.jsonl").read_text(encoding="utf-8").splitlines()
    ]

    assert observations[0]["fact_type"] == "percentage"
    assert observations[0]["normalized_value"] == "0.5"


def test_output_home_inside_input_is_rejected_before_writing(tmp_path: Path) -> None:
    bundle = tmp_path / "bundle"
    bundle.mkdir()
    project_root = Path(__file__).resolve().parents[1]
    home = bundle / ".proofmesh"
    service = DocumentAuditService(project_root, home)

    try:
        service.audit(str(bundle))
    except ValueError as exc:
        assert "不能放在待检查目录里面" in str(exc)
    else:
        raise AssertionError("expected output-home validation failure")
    assert not home.exists()
