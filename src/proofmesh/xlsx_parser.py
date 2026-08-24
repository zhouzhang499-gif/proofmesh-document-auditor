from __future__ import annotations

import hashlib
import re
from pathlib import Path
from typing import Any

import openpyxl
import yaml

from .contracts import EvidenceAtom, FactObservation
from .normalization import parse_fact


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def load_rules(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle)


def parse_xlsx(path: Path, root: Path, rules_path: Path) -> tuple[list[EvidenceAtom], list[FactObservation]]:
    rules = load_rules(rules_path)
    document_hash = sha256_file(path)
    relative_path = path.relative_to(root).as_posix()
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
    cached_workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    evidence: list[EvidenceAtom] = []
    observations: list[FactObservation] = []

    try:
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows()]
            cached_sheet = cached_workbook[sheet.title]
            sheet_money_unit = _infer_sheet_money_unit(rows)
            for row_index, row in enumerate(rows):
                for column_index, cell in enumerate(row):
                    if cell.value in (None, ""):
                        continue
                    locator = f"{sheet.title}!{cell.coordinate}"
                    evidence_id = hashlib.sha256(f"{document_hash}|{locator}".encode("utf-8")).hexdigest()[:20]
                    cached_value = cached_sheet[cell.coordinate].value if cell.data_type == "f" else cell.value
                    atom = EvidenceAtom(
                        evidence_id=evidence_id,
                        document_hash=document_hash,
                        relative_path=relative_path,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        locator=locator,
                        raw_text=(
                            f"formula={cell.value}; cached={cached_value}"
                            if cell.data_type == "f" and cached_value is not None
                            else str(cell.value)
                        ),
                        extractor=(
                            "openpyxl-formula-missing-cache"
                            if cell.data_type == "f" and cached_value is None
                            else "openpyxl"
                        ),
                    )
                    evidence.append(atom)
                    label = _infer_label(rows, row_index, column_index)
                    column_period = _infer_column_period(rows, row_index, column_index)
                    if column_period:
                        label = f"{label} {column_period}".strip()
                    fact_value = cached_value
                    if (
                        sheet_money_unit
                        and isinstance(fact_value, (int, float))
                        and "%" not in (cell.number_format or "")
                        and any(term in label for term in rules.get("money_labels", []))
                    ):
                        fact_value = f"{fact_value}{sheet_money_unit}"
                    if fact_value is None:
                        continue
                    fact = parse_fact(
                        value=fact_value,
                        label=label,
                        number_format=cell.number_format or "General",
                        document_hash=document_hash,
                        relative_path=relative_path,
                        locator=locator,
                        evidence_id=evidence_id,
                        aliases=rules.get("aliases", {}),
                        money_labels=rules.get("money_labels", []),
                        date_labels=rules.get("date_labels", []),
                    )
                    if fact:
                        observations.append(fact)
    finally:
        workbook.close()
        cached_workbook.close()

    return evidence, observations


def _infer_label(rows: list[list[Any]], row_index: int, column_index: int) -> str:
    row = rows[row_index]
    for index in range(column_index - 1, -1, -1):
        value = row[index].value
        if isinstance(value, str) and value.strip() and not _looks_numeric(value):
            return value.strip()
    for index in range(row_index - 1, -1, -1):
        value = rows[index][column_index].value if column_index < len(rows[index]) else None
        if isinstance(value, str) and value.strip() and not _looks_numeric(value):
            return value.strip()
    return ""


def _looks_numeric(value: str) -> bool:
    compact = value.replace(",", "").replace(".", "").replace("%", "").replace("％", "")
    for unit in ("亿元", "万元", "元", "人民币", "CNY"):
        compact = compact.replace(unit, "")
    return compact.strip().lstrip("+-").isdigit()


def _infer_column_period(rows: list[list[Any]], row_index: int, column_index: int) -> str:
    for index in range(row_index - 1, -1, -1):
        value = rows[index][column_index].value if column_index < len(rows[index]) else None
        if isinstance(value, str):
            match = re.search(r"20\d{2}(?:\s*[年-]?\s*[Qq][1-4])?", value)
            if match:
                return match.group(0)
        elif isinstance(value, int) and 2000 <= value <= 2099:
            return str(value)
    return ""


def _infer_sheet_money_unit(rows: list[list[Any]]) -> str:
    for row in rows[:10]:
        for cell in row[:10]:
            if isinstance(cell.value, str):
                match = re.search(r"单位\s*[:：]?\s*(亿元|万元|元|人民币|CNY)", cell.value, flags=re.IGNORECASE)
                if match:
                    return match.group(1)
    return ""
